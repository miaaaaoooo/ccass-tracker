"""
HKEX CCASS 持股量爬虫
====================
功能: 抓取港交所 CCASS 系统中指定股票 (默认 3836) 最近一周的 CCASS 参与者持股数据,
      输出为与模板一致的 Excel 文件,并自动计算 "近两日" 和 "近一周" 两列差值。

数据来源: https://www3.hkexnews.hk/sdw/search/searchsdw.aspx

依赖:
    pip install requests beautifulsoup4 openpyxl pandas lxml

用法:
    python hkex_ccass_crawler.py                 # 默认抓 3836 最近一周
    python hkex_ccass_crawler.py --stock 03836 --days 5
"""

import argparse
import datetime as dt
import re
import sys
import time
from typing import Dict, List, Optional

import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ----------------------------------------------------------------------
# 配置
# ----------------------------------------------------------------------
BASE_URL = "https://www3.hkexnews.hk/sdw/search/searchsdw.aspx"
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0 Safari/537.36"
    ),
    "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
}
# 香港公众假期(可按年补充);爬虫会跳过周末与此清单中的日期
HK_HOLIDAYS_2026 = {
    "2026-01-01", "2026-02-17", "2026-02-18", "2026-02-19",
    "2026-04-03", "2026-04-06", "2026-04-07", "2026-05-01",
    "2026-05-25", "2026-06-19", "2026-07-01", "2026-10-01",
    "2026-10-19", "2026-12-25",
}


# ----------------------------------------------------------------------
# 工具函数
# ----------------------------------------------------------------------
def previous_trading_days(end_date: dt.date, n: int) -> List[dt.date]:
    """从 end_date 往前推,返回最近 n 个港股交易日 (含 end_date,若它是交易日)."""
    days, cur = [], end_date
    while len(days) < n:
        if cur.weekday() < 5 and cur.strftime("%Y-%m-%d") not in HK_HOLIDAYS_2026:
            days.append(cur)
        cur -= dt.timedelta(days=1)
    return days


def parse_int(text: str) -> Optional[int]:
    """将 '524,795,660' 等带逗号的字符串解析成 int."""
    if not text:
        return None
    s = re.sub(r"[^\d-]", "", text)
    return int(s) if s else None


# ----------------------------------------------------------------------
# 核心爬虫
# ----------------------------------------------------------------------
class CCASSClient:
    """封装对 hkexnews.hk CCASS 查询页面的访问 (处理 ASP.NET __VIEWSTATE)."""

    def __init__(self):
        self.session = requests.Session()
        self.session.headers.update(HEADERS)
        self._form_state: Dict[str, str] = {}

    def _refresh_state(self):
        """GET 一次主页,抓取 __VIEWSTATE / __EVENTVALIDATION 等隐藏字段."""
        r = self.session.get(BASE_URL, timeout=20)
        r.raise_for_status()
        soup = BeautifulSoup(r.text, "lxml")
        self._form_state = {
            inp["name"]: inp.get("value", "")
            for inp in soup.select("input[type=hidden]")
            if inp.get("name")
        }

    def fetch(self, stock_code: str, query_date: dt.date) -> pd.DataFrame:
        """抓取指定日期、指定股票的 CCASS 持股明细,返回 DataFrame."""
        if not self._form_state:
            self._refresh_state()

        payload = {
            **self._form_state,
            "__EVENTTARGET": "btnSearch",
            "__EVENTARGUMENT": "",
            "today": dt.date.today().strftime("%Y%m%d"),
            "sortBy": "shareholding",
            "sortDirection": "desc",
            "txtShareholdingDate": query_date.strftime("%Y/%m/%d"),
            "txtStockCode": stock_code.zfill(5),     # 3836 -> 03836
            "txtStockName": "",
            "txtParticipantID": "",
            "txtParticipantName": "",
            "btnSearch.x": "1",
            "btnSearch.y": "1",
        }
        r = self.session.post(BASE_URL, data=payload, timeout=30)
        r.raise_for_status()
        return self._parse_table(r.text, query_date)

    @staticmethod
    def _parse_table(html: str, query_date: dt.date) -> pd.DataFrame:
        """解析持股明细表 (HKEX 用响应式 mobile-list 表格,每格里 heading+body 两段)."""
        soup = BeautifulSoup(html, "lxml")
        table = soup.find("table", class_="table-mobile-list")
        if table is None:
            raise RuntimeError(f"找不到持股明细表 ({query_date})")

        records = []
        for tr in table.select("tbody tr"):
            tds = tr.find_all("td")
            if len(tds) < 4:
                continue
            cells = []
            for td in tds:
                body = td.find("div", class_="mobile-list-body")
                cells.append(body.get_text(strip=True) if body else td.get_text(strip=True))
            # HKEX 列顺序: Participant ID | Name | Address | Shareholding | % of Issued
            records.append({
                "participant_id": cells[0],
                "name": re.sub(r"\s+", " ", cells[1]),
                "shareholding": parse_int(cells[3]),
                "pct_of_issued": cells[4] if len(cells) > 4 else None,
                "date": query_date,
            })
        if not records:
            raise RuntimeError(f"未解析到任何持股数据 ({query_date})")
        return pd.DataFrame(records)


# ----------------------------------------------------------------------
# 汇总 + 生成 Excel
# ----------------------------------------------------------------------
def build_pivot(daily: Dict[dt.date, pd.DataFrame]) -> pd.DataFrame:
    """把多日 DataFrame 透视成 (Participant x Date) 的宽表."""
    frames = []
    for d, df in daily.items():
        frames.append(df[["name", "shareholding"]].rename(columns={"shareholding": d}))
    wide = frames[0]
    for f in frames[1:]:
        wide = wide.merge(f, on="name", how="outer")
    # 按最新一天持股量降序
    latest = max(daily.keys())
    wide = wide.sort_values(by=latest, ascending=False, na_position="last").reset_index(drop=True)
    return wide


def write_excel(pivot: pd.DataFrame, dates: List[dt.date], stock_code: str, out_path: str):
    """按模板格式输出 xlsx:第一列序号,第二列名称,日期列 + 近两日/近一周差值列."""
    wb = Workbook()
    ws = wb.active
    ws.title = f"CCASS_{stock_code}"

    # 头部样式
    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="305496")
    diff_fill = PatternFill("solid", start_color="FFF2CC")
    center = Alignment(horizontal="center", vertical="center")

    # 列顺序: 序号 | 名称 | D0 | D1 | 近两日 | D2 | D3 | D4 | D5 | 近一周
    # dates 已按 [最近 -> 久远] 排序
    headers = ["排名", "Name of CCASS Participant",
               dates[0], dates[1], "近两日",
               dates[2], dates[3], dates[4], dates[5], "近一周"]
    ws.append(headers)

    # 写入数据行 (使用 Excel 公式而非硬编码差值)
    for idx, row in pivot.iterrows():
        excel_row = idx + 2  # Excel 第 2 行起
        values = [
            idx + 1,
            row["name"],
            row.get(dates[0]),
            row.get(dates[1]),
            f"=C{excel_row}-D{excel_row}",         # 近两日
            row.get(dates[2]),
            row.get(dates[3]),
            row.get(dates[4]),
            row.get(dates[5]),
            f"=C{excel_row}-I{excel_row}",         # 近一周
        ]
        ws.append(values)

    # 样式
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        if isinstance(h, dt.date):
            c.value = h
            c.number_format = "yyyy-mm-dd"

    for r in range(2, ws.max_row + 1):
        for diff_col in ("E", "J"):
            ws[f"{diff_col}{r}"].fill = diff_fill
            ws[f"{diff_col}{r}"].font = Font(name="Arial", bold=True)
        for col_idx in range(3, 11):
            ws.cell(row=r, column=col_idx).number_format = "#,##0;[Red](#,##0);-"

    # 列宽
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 42
    for col in range(3, 11):
        ws.column_dimensions[get_column_letter(col)].width = 16

    ws.freeze_panes = "C2"
    wb.save(out_path)
    print(f"已写出 {out_path}")


# ----------------------------------------------------------------------
# 入口
# ----------------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--stock", default="03836", help="股票代码 (默认 03836)")
    ap.add_argument("--days", type=int, default=6, help="抓取的交易日数 (默认 6)")
    ap.add_argument("--end", default=None, help="截止日期 YYYY-MM-DD,默认昨天")
    ap.add_argument("--out", default=None, help="输出文件名")
    args = ap.parse_args()

    end = (dt.date.fromisoformat(args.end) if args.end
           else dt.date.today() - dt.timedelta(days=1))
    trading_days = previous_trading_days(end, args.days)
    print(f"将抓取 {args.stock} 在以下交易日的 CCASS 持股: {[d.isoformat() for d in trading_days]}")

    client = CCASSClient()
    daily: Dict[dt.date, pd.DataFrame] = {}
    for d in trading_days:
        for attempt in range(3):
            try:
                daily[d] = client.fetch(args.stock, d)
                print(f"  ✓ {d}  {len(daily[d])} 条记录")
                time.sleep(1.0)            # 礼貌延时,避免被封
                break
            except Exception as e:
                print(f"  ! {d} 第 {attempt+1} 次失败: {e}", file=sys.stderr)
                time.sleep(3)
        else:
            print(f"  × {d} 放弃,该日数据缺失", file=sys.stderr)

    if len(daily) < 6:
        print("警告:抓到的交易日不足 6 天,差值列可能不完整。", file=sys.stderr)

    pivot = build_pivot(daily)
    out_path = args.out or f"CCASS_{args.stock}_{trading_days[0]}.xlsx"
    write_excel(pivot, trading_days, args.stock, out_path)


if __name__ == "__main__":
    main()

