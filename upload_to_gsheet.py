"""读取最新生成的 CCASS xlsx,推送到 Google Sheets。"""
import datetime as dt
import glob
import json
import os
import sys

import gspread
from google.oauth2.service_account import Credentials
from openpyxl import load_workbook

SPREADSHEET_ID = os.environ.get(
    "CCASS_SPREADSHEET_ID", "1lZwHxiF1_HlHshDy_gjvIAu3W_5U31gXiSB4_HQkamI")
# 凭证来源优先级: 环境变量里的 JSON 内容 > 环境变量指定的路径 > 同目录默认文件
CREDENTIALS_FILE = os.environ.get("GOOGLE_CREDENTIALS_FILE",
                                  "google_credentials.json")
WORKSHEET_NAME = "CCASS_03836"
XLSX_PATTERN = "CCASS_03836_*.xlsx"
SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]


def find_latest_xlsx():
    files = glob.glob(XLSX_PATTERN)
    if not files:
        sys.exit(f"找不到匹配 {XLSX_PATTERN} 的文件")
    latest = max(files, key=os.path.getmtime)
    print(f"使用最新文件: {latest}")
    return latest


def read_xlsx_as_rows(path):
    """读 xlsx,差值列在 Python 里现算 (不依赖 Excel 公式缓存)。"""
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = []
    for r in ws.iter_rows(values_only=True):
        new_row = []
        for v in r:
            if isinstance(v, dt.datetime):
                new_row.append(v.strftime("%Y-%m-%d"))
            elif v is None:
                new_row.append("")
            else:
                new_row.append(v)
        rows.append(new_row)

    # 重新计算差值列 E (近两日 = C - D) 和 J (近一周 = C - I)
    # 列索引 (0-based): A=0, B=1, C=2, D=3, E=4, F=5, G=6, H=7, I=8, J=9
    # 表头在第 3 行 (索引 2),数据从第 4 行 (索引 3) 开始
    for i, row in enumerate(rows):
        if i < 3:
            continue  # 跳过标题、副标题、表头
        if len(row) < 10:
            continue
        try:
            c = row[2] if isinstance(row[2], (int, float)) else None
            d = row[3] if isinstance(row[3], (int, float)) else None
            i_val = row[8] if isinstance(row[8], (int, float)) else None
            row[4] = (c - d) if (c is not None and d is not None) else ""
            row[9] = (c - i_val) if (c is not None and i_val is not None) else ""
        except Exception:
            pass
    return rows


def load_credentials():
    """CI 环境把密钥放在 GOOGLE_CREDENTIALS_JSON 里,本地则读文件。"""
    raw = os.environ.get("GOOGLE_CREDENTIALS_JSON")
    if raw:
        print("使用环境变量 GOOGLE_CREDENTIALS_JSON 中的凭证")
        return Credentials.from_service_account_info(json.loads(raw), scopes=SCOPES)
    if not os.path.exists(CREDENTIALS_FILE):
        sys.exit(f"找不到凭证: 既无 GOOGLE_CREDENTIALS_JSON 环境变量,也无 {CREDENTIALS_FILE}")
    print(f"使用凭证文件: {CREDENTIALS_FILE}")
    return Credentials.from_service_account_file(CREDENTIALS_FILE, scopes=SCOPES)


def push_to_sheets(rows):
    creds = load_credentials()
    gc = gspread.authorize(creds)
    sh = gc.open_by_key(SPREADSHEET_ID)
    try:
        ws = sh.worksheet(WORKSHEET_NAME)
        ws.clear()
        print(f"已清空已有 sheet: {WORKSHEET_NAME}")
    except gspread.WorksheetNotFound:
        ws = sh.add_worksheet(title=WORKSHEET_NAME, rows=200, cols=20)
        print(f"新建 sheet: {WORKSHEET_NAME}")

    stamp = [[f"最后更新: {dt.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"]]
    ws.update(values=stamp + [[""]] + rows, range_name="A1",
              value_input_option="USER_ENTERED")
    print(f"已写入 {len(rows)} 行数据 -> Google Sheets")

    # 表头格式
    ws.format("A3:J3", {
        "textFormat": {"bold": True, "foregroundColor":
                       {"red": 1, "green": 1, "blue": 1}},
        "backgroundColor": {"red": 0.19, "green": 0.33, "blue": 0.59},
        "horizontalAlignment": "CENTER",
    })
    # 差值列高亮
    ws.format("E4:E200", {"backgroundColor": {"red": 1, "green": 0.95, "blue": 0.8},
                          "textFormat": {"bold": True}})
    ws.format("J4:J200", {"backgroundColor": {"red": 1, "green": 0.95, "blue": 0.8},
                          "textFormat": {"bold": True}})
    # 数字格式 (带千分位,负数红色括号)
    ws.format("C4:J200", {"numberFormat":
              {"type": "NUMBER", "pattern": "#,##0;[Red](#,##0);-"}})


def main():
    xlsx = find_latest_xlsx()
    rows = read_xlsx_as_rows(xlsx)
    push_to_sheets(rows)
    print("OK 完成")


if __name__ == "__main__":
    main()
    